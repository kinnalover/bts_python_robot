import sqlite3

from openpyxl import load_workbook

import pyexcel as p


def load_xlsx_to_db(filepath):
    print("Начинаем обработку excel файла")
    try:
        file_format = filepath.split(".")[-1]
        if file_format == "xls":
            p.save_book_as(file_name=filepath,
                           dest_file_name=filepath + "x")
            filepath += "x"
        wb = load_workbook(filename=filepath)
        sheet_name = ""

        if len(wb.sheetnames) > 1:
            for sheet in wb.sheetnames:
                if "рус" in sheet:
                    sheet_name = sheet
                elif "Лист1" in sheet:
                    sheet_name = sheet
                else:
                    count = 0
                    for i in range(len(wb.sheetnames)):
                        if len(wb[wb.sheetnames[i]]['A']) > count:
                            count = len(wb[wb.sheetnames[i]]['A'])
                            sheet_name = wb.sheetnames[i]


        else:
            sheet_name = wb.sheetnames[0]

        my_sheet = wb[sheet_name]

        for i in range(1, len(my_sheet['B'])):

            if my_sheet.cell(row=i, column=2).value is not None:
                my_len = len(str(my_sheet.cell(row=i, column=2).value))
                if my_len == 12:
                    bin = str(my_sheet.cell(row=i, column=2).value)
                    fio = str(my_sheet.cell(row=i, column=3).value)
                    gov_reg = str(my_sheet.cell(row=i, column=4).value)
                    address = str(my_sheet.cell(row=i, column=5).value)
                    date_definition = my_sheet.cell(row=i, column=6).value
                    date_nomination = str(my_sheet.cell(row=i, column=7).value)
                    date_temp_manager = str(my_sheet.cell(row=i, column=8).value)
                    fio_temp_manager = str(my_sheet.cell(row=i, column=9).value)
                    date_accept_from = str(my_sheet.cell(row=i, column=10).value)
                    date_accept_till = str(my_sheet.cell(row=i, column=11).value)
                    address_reception = str(my_sheet.cell(row=i, column=12).value)
                    contact_number = str(my_sheet.cell(row=i, column=13).value)
                    date_registration = str(my_sheet.cell(row=i, column=14).value)
                    info_list = [bin, fio, gov_reg, address, date_definition, date_nomination, date_temp_manager,
                                 fio_temp_manager, date_accept_from, date_accept_till, address_reception,
                                 contact_number,
                                 date_registration]

                    bool_insert = check_if_already_has_row(info_list[0], info_list[2], info_list[12])
                    if bool_insert:
                        insert_db(info_list)
        return True

    except Exception as ex:
        print(ex)
        return False


def insert_db(info_list):
    conn = sqlite3.connect('companies.db')
    cur = conn.cursor()
    try:
        cur.execute(
            f"INSERT OR REPLACE INTO companies (bin,fio,gov_reg,address,date_definition,date_nomination,date_temp_manager,fio_temp_manager,date_accept_from,date_accept_till,address_reception,contact_number,date_registration) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?);",
            (info_list[0], info_list[1], info_list[2], info_list[3], info_list[4], info_list[5], info_list[6],
             info_list[7],
             info_list[8], info_list[9], info_list[10], info_list[11], info_list[12]))
        conn.commit()
        return True
    except Exception as ex:
        print(ex)
        return False


def check_if_already_has_row(bin, gov_reg, reg):
    conn = sqlite3.connect('companies.db')
    cur = conn.cursor()
    try:
        cur.execute("Select * from companies where bin='" + bin + "'")
        result = cur.fetchall()
        if len(result) >= 1:
            for item in result:
                if item[2] == gov_reg and item[12] == reg:
                    return False
        else:
            return True
    except Exception as ex:
        print(ex)
        print("check_if_already_has_row")

        return False
